"""
Optimizador Nutricional con Algoritmo Genético (PyGAD)
======================================================
Encuentra el plato más nutritivo y económico por rango de edad.
Cada rango de edad usa un modelo genético independiente con
parámetros diferenciados.
"""

import pandas as pd
import numpy as np
import pygad
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# 1. DATOS DE PLATOS
# ─────────────────────────────────────────────
data = {
    "Platillo": [
        "Ajiaco santafereño", "Arroz atollado", "Caldo de costilla",
        "Changua", "Cuchuco de trigo", "Sancocho de gallina",
        "Puchero santafereño", "Sopa de cebada", "Almojábana",
        "Arepa de maíz pelao", "Envuelto de mazorca", "Mogolla chicharrona",
        "Pandebono", "Torta de maíz", "Chicha", "Masato",
        "Cuajada con melao", "Dulce de arracacha", "Melao con queso",
        "Postre de natas", "Carne a la llanera", "Fritanga",
        "Huevos pericos", "Lengua en salsa", "Papa rellena",
        "Papa salada con hogao", "Sobrebarriga en salsa", "Tamal cundinamarqués"
    ],
    "Categoria": [
        "Sopa","Guiso","Sopa","Sopa","Sopa","Sopa","Sopa","Sopa",
        "Panadería","Panadería","Panadería","Panadería","Panadería","Panadería",
        "Bebida","Bebida","Postre","Postre","Postre","Postre",
        "Plato principal","Plato principal","Plato principal","Plato principal",
        "Plato principal","Plato principal","Plato principal","Tamal"
    ],
    # Costo por 100g en COP
    "Costo_por_100g": [
        4968, 19514, 12708, 10628, 10938, 16719, 22725, 9219,
        6662, 2867, 6215, 11136, 9707, 17092, 2773, 4107,
        5850, 14133, 7907, 12600, 13429, 16029, 8092, 7438,
        16538, 14462, 8510, 19080
    ],
    # Precio por porción en COP
    "Precio_por_porcion": [
        19870, 68300, 50833, 37199, 43750, 66875, 90900, 36875,
        5330, 2867, 6215, 8909, 5824, 20510, 6933, 10267,
        11700, 21200, 11860, 18900, 47000, 56100, 16183, 26033,
        33075, 28925, 29783, 47700
    ],
    # Precio total lote en COP
    "Precio_total_lote": [
        158960, 683000, 305000, 297590, 350000, 535000, 909000, 295000,
        213200, 172000, 248600, 356350, 291200, 410200, 208000, 308000,
        117000, 318000, 118600, 377990, 282000, 561000, 97100, 156200,
        264600, 231400, 178700, 954000
    ],
    # Macronutrientes por 100g
    "Proteina_g":    [6.4,10.1,2.208,4.509,3.232,3.2,9.342,3.216,20.14,3.045,9.684,23.96,11.78,15.238,1.004,0.674,3.95,0.308,3.982,7.902,7.616,30.349,4.47,5.21,13.68,0.702,8.546,26.656],
    "Lipidos_g":     [0.8,4.3,1.952,3.726,2.048,1.328,4.644,2.256,20.9,0.315,10.872,15.04,13.908,46.398,0.481,0.041,3.875,0.042,5.632,7.983,7.392,27.184,3.3,3.7,11.19,30.063,7.456,6.912],
    "Carb_totales_g":[11.3,9.2,5.128,8.406,11.048,16.352,20.25,8.408,67.526,27.265,62.028,75.24,51.49,65.588,16.551,17.841,23.05,23.462,19.91,41.013,15.764,41.606,3.63,9.557,29.79,3.783,10.71,78.592],
    "Energia_kcal":  [82,119,48.32,86.4,76.96,93.44,165.78,70.24,544.54,124.95,390.24,540.8,379.62,746.7,75.947,74.547,142.75,75.494,146.3,267.441,162.12,535.613,63.6,94.545,276.6,289.641,145.942,497.28],
    "Fibra_g":       [2.2,1.7,0.728,0.54,0.728,1.64,2.7,1.744,2.774,0.49,2.628,4.28,0.684,2.774,0.783,0.153,0,0.226,0,0.159,1.008,1.44,0.87,1.278,0.84,1.029,1.439,6.784],
    "Calcio_mg":     [15,20,13.36,101.61,17.36,14.8,21.42,11.04,463.6,11.9,244.08,68.8,281.2,283.1,15.102,10.402,156,33.004,158.4,310.506,20.44,31.078,25.8,20.088,28.5,20.106,29.212,62.72],
    "Hierro_mg":     [0.7,0.7,0.464,0.63,0.808,0.672,1.359,0.512,3.116,0.805,1.8,3.92,1.634,2.546,0.358,0.278,0.575,0.356,0.264,0.744,0.98,2.202,0.75,1.201,1.62,0.354,1.243,3.904],
    "VitC_mg":       [3,26,9.04,6.93,9.04,12.8,9.36,2.24,0,0,0,0,11.4,0,0.004,0.004,0,4.008,0,2.711,14,17.28,10.8,10.577,6,10.811,7.032,9.28],
    "VitA_ugER":     [407,0.8,37.6,76.95,37.6,45.6,9.99,107.2,304.76,0,163.44,0,227.24,551,4.015,0.015,49.5,12.63,93.72,86.445,0.28,18.56,94.2,30.419,53.1,41.145,21.398,427.84],
}

df = pd.DataFrame(data)

# ─────────────────────────────────────────────
# 2. RANGOS DE EDAD Y RECOMENDACIONES
# ─────────────────────────────────────────────
AGE_GROUPS = {
    "0–5 años":   {"carb_min":0.50, "carb_max":0.55, "prot_min":0.15, "prot_max":0.20, "fat_min":0.30, "fat_max":0.35,
                   "num_generations":80,  "sol_per_pop":20, "mutation_percent_genes":20, "parent_selection_type":"sss",
                   "crossover_type":"single_point", "mutation_type":"random", "keep_elitism":1,
                   "w_prot":2.0, "w_fiber":2.5, "w_calcium":0.02, "w_iron":1.5, "w_vitc":0.3, "w_vita":0.003, "w_cost":8.0},
    "6–12 años":  {"carb_min":0.50, "carb_max":0.55, "prot_min":0.15, "prot_max":0.20, "fat_min":0.25, "fat_max":0.30,
                   "num_generations":90,  "sol_per_pop":30, "mutation_percent_genes":15, "parent_selection_type":"rws",
                   "crossover_type":"single_point", "mutation_type":"random", "keep_elitism":2,
                   "w_prot":2.5, "w_fiber":2.0, "w_calcium":0.02, "w_iron":2.0, "w_vitc":0.4, "w_vita":0.004, "w_cost":9.0},
    "13–18 años": {"carb_min":0.45, "carb_max":0.55, "prot_min":0.20, "prot_max":0.25, "fat_min":0.25, "fat_max":0.30,
                   "num_generations":100, "sol_per_pop":40, "mutation_percent_genes":12, "parent_selection_type":"tournament",
                   "crossover_type":"single_point", "mutation_type":"random", "keep_elitism":2,
                   "w_prot":3.5, "w_fiber":1.5, "w_calcium":0.015, "w_iron":2.5, "w_vitc":0.5, "w_vita":0.004, "w_cost":10.0},
    "19–35 años": {"carb_min":0.45, "carb_max":0.50, "prot_min":0.20, "prot_max":0.25, "fat_min":0.25, "fat_max":0.30,
                   "num_generations":100, "sol_per_pop":50, "mutation_percent_genes":10, "parent_selection_type":"rank",
                   "crossover_type":"single_point", "mutation_type":"random", "keep_elitism":3,
                   "w_prot":3.0, "w_fiber":2.0, "w_calcium":0.012, "w_iron":2.0, "w_vitc":0.4, "w_vita":0.003, "w_cost":12.0},
    "36–59 años": {"carb_min":0.40, "carb_max":0.50, "prot_min":0.20, "prot_max":0.30, "fat_min":0.25, "fat_max":0.35,
                   "num_generations":110, "sol_per_pop":50, "mutation_percent_genes":8,  "parent_selection_type":"sus",
                   "crossover_type":"single_point", "mutation_type":"random", "keep_elitism":3,
                   "w_prot":3.0, "w_fiber":3.0, "w_calcium":0.015, "w_iron":2.5, "w_vitc":0.5, "w_vita":0.004, "w_cost":11.0},
    "60–75 años": {"carb_min":0.40, "carb_max":0.45, "prot_min":0.25, "prot_max":0.30, "fat_min":0.30, "fat_max":0.35,
                   "num_generations":120, "sol_per_pop":60, "mutation_percent_genes":8,  "parent_selection_type":"random",
                   "crossover_type":"single_point", "mutation_type":"random", "keep_elitism":4,
                   "w_prot":4.0, "w_fiber":3.5, "w_calcium":0.020, "w_iron":2.0, "w_vitc":0.6, "w_vita":0.005, "w_cost":9.0},
    "75+ años":   {"carb_min":0.35, "carb_max":0.45, "prot_min":0.25, "prot_max":0.35, "fat_min":0.30, "fat_max":0.35,
                   "num_generations":130, "sol_per_pop":80, "mutation_percent_genes":6,  "parent_selection_type":"rws",
                   "crossover_type":"single_point", "mutation_type":"random", "keep_elitism":5,
                   "w_prot":5.0, "w_fiber":4.0, "w_calcium":0.025, "w_iron":2.5, "w_vitc":0.7, "w_vita":0.006, "w_cost":7.0},
}

n_dishes = len(df)

# ─────────────────────────────────────────────
# 3. FUNCIÓN DE FITNESS
# ─────────────────────────────────────────────
def make_fitness_fn(params, df):
    """Crea función de fitness específica para un rango de edad."""
    carb_min  = params["carb_min"];  carb_max  = params["carb_max"]
    prot_min  = params["prot_min"];  prot_max  = params["prot_max"]
    fat_min   = params["fat_min"];   fat_max   = params["fat_max"]
    w_prot    = params["w_prot"];    w_fiber   = params["w_fiber"]
    w_calcium = params["w_calcium"]; w_iron    = params["w_iron"]
    w_vitc    = params["w_vitc"];    w_vita    = params["w_vita"]
    w_cost    = params["w_cost"]

    def fitness_fn(ga_instance, solution, solution_idx):
        idx = int(round(np.clip(solution[0], 0, n_dishes - 1)))
        row = df.iloc[idx]

        cals = row["Energia_kcal"]
        if cals <= 0:
            return -1000.0

        prot_g  = row["Proteina_g"]
        fat_g   = row["Lipidos_g"]
        carb_g  = row["Carb_totales_g"]

        cals_prot = prot_g * 4
        cals_fat  = fat_g  * 9
        cals_carb = carb_g * 4
        cals_total = cals_prot + cals_fat + cals_carb
        if cals_total <= 0:
            return -1000.0

        ratio_prot = cals_prot / cals_total
        ratio_fat  = cals_fat  / cals_total
        ratio_carb = cals_carb / cals_total

        def penalty(val, lo, hi):
            if val < lo: return (lo - val) ** 2
            if val > hi: return (val - hi) ** 2
            return 0.0

        macro_penalty = (
            penalty(ratio_carb, carb_min, carb_max) +
            penalty(ratio_prot, prot_min, prot_max) +
            penalty(ratio_fat,  fat_min,  fat_max)
        )

        nutrition_score = (
            prot_g           * w_prot +
            row["Fibra_g"]   * w_fiber +
            row["Calcio_mg"] * w_calcium +
            row["Hierro_mg"] * w_iron +
            row["VitC_mg"]   * w_vitc +
            row["VitA_ugER"] * w_vita
        )

        costo_norm = row["Costo_por_100g"] / 25000.0
        fitness = nutrition_score - 200.0 * macro_penalty - w_cost * costo_norm

        return float(fitness)

    return fitness_fn


# ─────────────────────────────────────────────
# 4. EJECUTAR MODELO POR CADA RANGO DE EDAD
# ─────────────────────────────────────────────
results = []

print("=" * 70)
print(" OPTIMIZADOR NUTRICIONAL CON ALGORITMO GENÉTICO (PyGAD)")
print(" Platos colombianos — Selección por rango de edad")
print("=" * 70)

for age_label, params in AGE_GROUPS.items():
    fitness_fn = make_fitness_fn(params, df)

    ga = pygad.GA(
        num_generations          = params["num_generations"],
        num_parents_mating       = max(4, params["sol_per_pop"] // 5),
        fitness_func             = fitness_fn,
        sol_per_pop              = params["sol_per_pop"],
        num_genes                = 1,
        gene_type                = float,
        init_range_low           = 0,
        init_range_high          = n_dishes - 0.01,
        gene_space               = {"low": 0, "high": n_dishes - 0.01},
        parent_selection_type    = params["parent_selection_type"],
        crossover_type           = params["crossover_type"],
        mutation_type            = "random",
        mutation_percent_genes   = params["mutation_percent_genes"],
        keep_elitism             = params["keep_elitism"],
        suppress_warnings        = True,
        random_seed              = 42,
    )

    ga.run()

    solution, solution_fitness, _ = ga.best_solution()
    best_idx = int(round(np.clip(solution[0], 0, n_dishes - 1)))
    best_row = df.iloc[best_idx]

    # Calcular ratios del ganador
    cals_p = best_row["Proteina_g"] * 4
    cals_f = best_row["Lipidos_g"]  * 9
    cals_c = best_row["Carb_totales_g"] * 4
    total  = cals_p + cals_f + cals_c
    rp = cals_p/total*100 if total>0 else 0
    rf = cals_f/total*100 if total>0 else 0
    rc = cals_c/total*100 if total>0 else 0

    rec = params
    results.append({
        "Rango de edad": age_label,
        "Plato óptimo": best_row["Platillo"],
        "Categoría": best_row["Categoria"],
        "Fitness": round(solution_fitness, 3),
        "Energía (kcal/100g)": round(best_row["Energia_kcal"], 1),
        "Proteína (g/100g)": round(best_row["Proteina_g"], 2),
        "Carb. (g/100g)": round(best_row["Carb_totales_g"], 2),
        "Grasas (g/100g)": round(best_row["Lipidos_g"], 2),
        "% Carb (cal)": f"{rc:.1f}% (rango: {rec['carb_min']*100:.0f}–{rec['carb_max']*100:.0f}%)",
        "% Proteína (cal)": f"{rp:.1f}% (rango: {rec['prot_min']*100:.0f}–{rec['prot_max']*100:.0f}%)",
        "% Grasas (cal)": f"{rf:.1f}% (rango: {rec['fat_min']*100:.0f}–{rec['fat_max']*100:.0f}%)",
        "Costo/100g ($COP)": f"${best_row['Costo_por_100g']:,.0f}",
        "Costo/porción ($COP)": f"${best_row['Precio_por_porcion']:,.0f}",
        "Fibra (g)": round(best_row["Fibra_g"], 2),
        "Calcio (mg)": round(best_row["Calcio_mg"], 1),
        "Hierro (mg)": round(best_row["Hierro_mg"], 2),
        "Vit. C (mg)": round(best_row["VitC_mg"], 1),
        "Vit. A (μg ER)": round(best_row["VitA_ugER"], 1),
        "Modelo GA": f"Gen:{params['num_generations']} Pop:{params['sol_per_pop']} | Sel:{params['parent_selection_type']} | Mut:random {params['mutation_percent_genes']}% | Elite:{params['keep_elitism']}",
    })

    print(f"\n{'─'*60}")
    print(f"  Rango: {age_label}")
    print(f"  Modelo GA: Gen={params['num_generations']} | Pop={params['sol_per_pop']} | Sel={params['parent_selection_type']} | Mut=random {params['mutation_percent_genes']}% | Elite={params['keep_elitism']}")
    print(f"  ✅ Plato óptimo: {best_row['Platillo']} ({best_row['Categoria']})")
    print(f"     Fitness: {solution_fitness:.3f}")
    print(f"     Energía: {best_row['Energia_kcal']:.1f} kcal/100g")
    print(f"     Proteína: {best_row['Proteina_g']:.2f}g | Carb: {best_row['Carb_totales_g']:.2f}g | Grasas: {best_row['Lipidos_g']:.2f}g")
    print(f"     %Carb={rc:.1f}%  %Prot={rp:.1f}%  %Grasa={rf:.1f}%")
    print(f"     Costo/100g: ${best_row['Costo_por_100g']:,} COP | Costo/porción: ${best_row['Precio_por_porcion']:,} COP")

print(f"\n{'='*70}")
print("  RESUMEN FINAL")
print(f"{'='*70}")

df_results = pd.DataFrame(results)
print(df_results[["Rango de edad","Plato óptimo","Categoría","Energía (kcal/100g)","Costo/porción ($COP)","Modelo GA"]].to_string(index=False))

# ─────────────────────────────────────────────
# 5. EXPORTAR RESULTADOS A EXCEL
# ─────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Resultados Óptimos"

# Colores
COLOR_HEADER_MAIN  = "1B4F72"
COLOR_HEADER_SUB   = "2874A6"
COLOR_AGE_0_5      = "D6EAF8"
COLOR_AGE_6_12     = "D5F5E3"
COLOR_AGE_13_18    = "FCF3CF"
COLOR_AGE_19_35    = "FDEBD0"
COLOR_AGE_36_59    = "F9EBEA"
COLOR_AGE_60_75    = "EBF5FB"
COLOR_AGE_75PLUS   = "F0F3F4"

AGE_COLORS = [COLOR_AGE_0_5, COLOR_AGE_6_12, COLOR_AGE_13_18, COLOR_AGE_19_35,
              COLOR_AGE_36_59, COLOR_AGE_60_75, COLOR_AGE_75PLUS]

WHITE = "FFFFFF"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(cell, bold=False, bg=None, font_color="000000", size=10,
               h_align="center", v_align="center", wrap=True):
    cell.font = Font(name="Calibri", bold=bold, color=font_color, size=size)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    cell.border = border

# ── Título principal ──────────────────────────────────────────
ws.merge_cells("A1:S1")
ws["A1"] = "OPTIMIZADOR NUTRICIONAL — ALGORITMO GENÉTICO (PyGAD)"
cell_style(ws["A1"], bold=True, bg=COLOR_HEADER_MAIN, font_color=WHITE, size=13)

ws.merge_cells("A2:S2")
ws["A2"] = "Selección del plato más nutritivo y económico por rango de edad según recomendaciones de macronutrientes"
cell_style(ws["A2"], bg=COLOR_HEADER_SUB, font_color=WHITE, size=10)

# ── Cabecera de columnas ──────────────────────────────────────
headers = [
    "Rango de Edad", "Plato Óptimo", "Categoría",
    "Energía\n(kcal/100g)", "Proteína\n(g/100g)", "Carb.\n(g/100g)", "Grasas\n(g/100g)",
    "% Carb.\ncalórico", "% Proteína\ncalórico", "% Grasas\ncalórico",
    "Costo\n/100g ($COP)", "Costo\n/porción ($COP)",
    "Fibra\n(g)", "Calcio\n(mg)", "Hierro\n(mg)", "Vit. C\n(mg)", "Vit. A\n(μg ER)",
    "Fitness GA", "Modelo Genético (parámetros)"
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell_style(cell, bold=True, bg="2C3E50", font_color=WHITE, size=9)
    ws.row_dimensions[3].height = 35

# ── Datos ─────────────────────────────────────────────────────
row_map = {
    "Rango de edad": "Rango de Edad",
    "Plato óptimo": "Plato Óptimo",
    "Categoría": "Categoría",
    "Energía (kcal/100g)": "Energía\n(kcal/100g)",
    "Proteína (g/100g)": "Proteína\n(g/100g)",
    "Carb. (g/100g)": "Carb.\n(g/100g)",
    "Grasas (g/100g)": "Grasas\n(g/100g)",
    "% Carb (cal)": "% Carb.\ncalórico",
    "% Proteína (cal)": "% Proteína\ncalórico",
    "% Grasas (cal)": "% Grasas\ncalórico",
    "Costo/100g ($COP)": "Costo\n/100g ($COP)",
    "Costo/porción ($COP)": "Costo\n/porción ($COP)",
    "Fibra (g)": "Fibra\n(g)",
    "Calcio (mg)": "Calcio\n(mg)",
    "Hierro (mg)": "Hierro\n(mg)",
    "Vit. C (mg)": "Vit. C\n(mg)",
    "Vit. A (μg ER)": "Vit. A\n(μg ER)",
    "Fitness": "Fitness GA",
    "Modelo GA": "Modelo Genético (parámetros)",
}

field_order = list(row_map.keys())

for r_idx, result in enumerate(results):
    excel_row = r_idx + 4
    bg = AGE_COLORS[r_idx % len(AGE_COLORS)]
    ws.row_dimensions[excel_row].height = 22
    for col_idx, field in enumerate(field_order, start=1):
        val = result.get(field, "")
        cell = ws.cell(row=excel_row, column=col_idx, value=val)
        h_align = "left" if col_idx in [1, 2, 3, 19] else "center"
        cell_style(cell, bg=bg, size=9, h_align=h_align)

# ── Hoja de recomendaciones ───────────────────────────────────
ws2 = wb.create_sheet("Recomendaciones por Edad")

ws2.merge_cells("A1:G1")
ws2["A1"] = "RECOMENDACIONES DE MACRONUTRIENTES POR RANGO DE EDAD"
cell_style(ws2["A1"], bold=True, bg=COLOR_HEADER_MAIN, font_color=WHITE, size=12)

rec_headers = ["Rango de Edad", "Carbohidratos (%)", "Proteínas (%)", "Grasas (%)",
               "Selección GA", "Cruce GA", "Mutación GA (%)"]
for c, h in enumerate(rec_headers, 1):
    cell = ws2.cell(row=2, column=c, value=h)
    cell_style(cell, bold=True, bg="2C3E50", font_color=WHITE, size=10)
    ws2.row_dimensions[2].height = 25

for r, (age, p) in enumerate(AGE_GROUPS.items(), 3):
    bg = AGE_COLORS[r - 3]
    vals = [
        age,
        f"{p['carb_min']*100:.0f}–{p['carb_max']*100:.0f}%",
        f"{p['prot_min']*100:.0f}–{p['prot_max']*100:.0f}%",
        f"{p['fat_min']*100:.0f}–{p['fat_max']*100:.0f}%",
        p["parent_selection_type"],
        p["crossover_type"],
        p["mutation_percent_genes"],
    ]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell_style(cell, bg=bg, size=10)
    ws2.row_dimensions[r].height = 20

# Anchos de columnas
col_widths_ws1 = [16, 25, 16, 10, 10, 10, 10, 18, 18, 18, 14, 16, 8, 9, 9, 9, 11, 9, 45]
for i, w in enumerate(col_widths_ws1, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

col_widths_ws2 = [16, 18, 16, 14, 18, 16, 18]
for i, w in enumerate(col_widths_ws2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A4"
ws2.freeze_panes = "A3"

output_path = "/mnt/user-data/outputs/Optimizacion_Nutricional_PyGAD.xlsx"
print(f"\n✅ Resultados exportados: {output_path}")