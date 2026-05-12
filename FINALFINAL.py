"""
Ranking Nutricional Completo por Rango de Edad — PyGAD
======================================================
Evalúa TODOS los platos con la función de fitness de cada modelo genético
y genera un ranking ordenado. Incorpora los 3 indicadores de costo:
  - Precio por porción
  - Costo por 100g
  - Precio total por lote
"""

import pandas as pd
import numpy as np
import pygad
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# 1. DATOS
# ─────────────────────────────────────────────────────────────────────────────
data = {
    "Platillo": [
        "Ajiaco santafereño","Arroz atollado","Caldo de costilla","Changua",
        "Cuchuco de trigo","Sancocho de gallina","Puchero santafereño","Sopa de cebada",
        "Almojábana","Arepa de maíz pelao","Envuelto de mazorca","Mogolla chicharrona",
        "Pandebono","Torta de maíz","Chicha","Masato",
        "Cuajada con melao","Dulce de arracacha","Melao con queso","Postre de natas",
        "Carne a la llanera","Fritanga","Huevos pericos","Lengua en salsa",
        "Papa rellena","Papa salada con hogao","Sobrebarriga en salsa","Tamal cundinamarqués"
    ],
    "Categoria": [
        "Sopa","Guiso","Sopa","Sopa","Sopa","Sopa","Sopa","Sopa",
        "Panadería","Panadería","Panadería","Panadería","Panadería","Panadería",
        "Bebida","Bebida","Postre","Postre","Postre","Postre",
        "Plato principal","Plato principal","Plato principal","Plato principal",
        "Plato principal","Plato principal","Plato principal","Tamal"
    ],
    "Costo_por_100g":    [4968,19514,12708,10628,10938,16719,22725,9219,6662,2867,6215,11136,9707,17092,2773,4107,5850,14133,7907,12600,13429,16029,8092,7438,16538,14462,8510,19080],
    "Precio_por_porcion":[19870,68300,50833,37199,43750,66875,90900,36875,5330,2867,6215,8909,5824,20510,6933,10267,11700,21200,11860,18900,47000,56100,16183,26033,33075,28925,29783,47700],
    "Precio_total_lote": [158960,683000,305000,297590,350000,535000,909000,295000,213200,172000,248600,356350,291200,410200,208000,308000,117000,318000,118600,377990,282000,561000,97100,156200,264600,231400,178700,954000],
    "Proteina_g":    [6.4,10.1,2.208,4.509,3.232,3.2,9.342,3.216,20.14,3.045,9.684,23.96,11.78,15.238,1.004,0.674,3.95,0.308,3.982,7.902,7.616,30.349,4.47,5.21,13.68,0.702,8.546,26.656],
    "Lipidos_g":     [0.8,4.3,1.952,3.726,2.048,1.328,4.644,2.256,20.9,0.315,10.872,15.04,13.908,46.398,0.481,0.041,3.875,0.042,5.632,7.983,7.392,27.184,3.3,3.7,11.19,30.063,7.456,6.912],
    "Carb_totales_g":[11.3,9.2,5.128,8.406,11.048,16.352,20.25,8.408,67.526,27.265,62.028,75.24,51.49,65.588,16.551,17.841,23.05,23.462,19.91,41.013,15.764,41.606,3.63,9.557,29.79,3.783,10.71,78.592],
    "Energia_kcal":  [82,119,48.32,86.4,76.96,93.44,165.78,70.24,544.54,124.95,390.24,540.8,379.62,746.7,75.947,74.547,142.75,75.494,146.3,267.441,162.12,535.613,63.6,94.545,276.6,289.641,145.942,497.28],
    "Fibra_g":       [2.2,1.7,0.728,0.54,0.728,1.64,2.7,1.744,2.774,0.49,2.628,4.28,0.684,2.774,0.783,0.153,0,0.226,0,0.159,1.008,1.44,0.87,1.278,0.84,1.029,1.439,6.784],
    "Calcio_mg":     [15,20,13.36,101.61,17.36,14.8,21.42,11.04,463.6,11.9,244.08,68.8,281.2,283.1,15.102,10.402,156,33.004,158.4,310.506,20.44,31.078,25.8,20.088,28.5,20.106,29.212,62.72],
    "Hierro_mg":     [0.7,0.7,0.464,0.63,0.808,0.672,1.359,0.512,3.116,0.805,1.8,3.92,1.634,2.546,0.358,0.278,0.575,0.356,0.264,0.744,0.98,2.202,0.75,1.201,1.62,0.354,1.243,3.904],
    "VitC_mg":       [3,26,9.04,6.93,9.04,12.8,9.36,2.24,0,0,0,0,11.4,0,0.004,0.004,0,4.008,0,2.711,14,17.28,10.8,10.577,6,10.811,7.032,9.28],
    "VitA_ugER":     [407,0.8,37.6,76.95,37.6,45.6,9.99,107.2,304.76,0,163.44,0,227.24,551,4.015,0.015,49.5,12.63,93.72,86.445,0.28,18.56,94.2,30.419,53.1,41.145,21.398,427.84],
}
df = pd.DataFrame(data)
n_dishes = len(df)

# ─────────────────────────────────────────────────────────────────────────────
# 2. CONFIGURACIÓN DE MODELOS GENÉTICOS POR EDAD
# ─────────────────────────────────────────────────────────────────────────────
AGE_GROUPS = {
    "0–5 años":   {"carb_min":0.50,"carb_max":0.55,"prot_min":0.15,"prot_max":0.20,"fat_min":0.30,"fat_max":0.35,
                   "num_generations":80, "sol_per_pop":20,"mutation_percent_genes":20,"parent_selection_type":"sss",
                   "keep_elitism":1,"w_prot":2.0,"w_fiber":2.5,"w_calcium":0.02,"w_iron":1.5,"w_vitc":0.3,"w_vita":0.003,
                   "w_cost100":6.0,"w_cost_porcion":3.0,"w_cost_lote":1.0},
    "6–12 años":  {"carb_min":0.50,"carb_max":0.55,"prot_min":0.15,"prot_max":0.20,"fat_min":0.25,"fat_max":0.30,
                   "num_generations":90, "sol_per_pop":30,"mutation_percent_genes":15,"parent_selection_type":"rws",
                   "keep_elitism":2,"w_prot":2.5,"w_fiber":2.0,"w_calcium":0.02,"w_iron":2.0,"w_vitc":0.4,"w_vita":0.004,
                   "w_cost100":6.0,"w_cost_porcion":3.5,"w_cost_lote":1.0},
    "13–18 años": {"carb_min":0.45,"carb_max":0.55,"prot_min":0.20,"prot_max":0.25,"fat_min":0.25,"fat_max":0.30,
                   "num_generations":100,"sol_per_pop":40,"mutation_percent_genes":12,"parent_selection_type":"tournament",
                   "keep_elitism":2,"w_prot":3.5,"w_fiber":1.5,"w_calcium":0.015,"w_iron":2.5,"w_vitc":0.5,"w_vita":0.004,
                   "w_cost100":7.0,"w_cost_porcion":4.0,"w_cost_lote":1.5},
    "19–35 años": {"carb_min":0.45,"carb_max":0.50,"prot_min":0.20,"prot_max":0.25,"fat_min":0.25,"fat_max":0.30,
                   "num_generations":100,"sol_per_pop":50,"mutation_percent_genes":10,"parent_selection_type":"rank",
                   "keep_elitism":3,"w_prot":3.0,"w_fiber":2.0,"w_calcium":0.012,"w_iron":2.0,"w_vitc":0.4,"w_vita":0.003,
                   "w_cost100":8.0,"w_cost_porcion":4.0,"w_cost_lote":2.0},
    "36–59 años": {"carb_min":0.40,"carb_max":0.50,"prot_min":0.20,"prot_max":0.30,"fat_min":0.25,"fat_max":0.35,
                   "num_generations":110,"sol_per_pop":50,"mutation_percent_genes":8, "parent_selection_type":"sus",
                   "keep_elitism":3,"w_prot":3.0,"w_fiber":3.0,"w_calcium":0.015,"w_iron":2.5,"w_vitc":0.5,"w_vita":0.004,
                   "w_cost100":8.0,"w_cost_porcion":4.5,"w_cost_lote":2.0},
    "60–75 años": {"carb_min":0.40,"carb_max":0.45,"prot_min":0.25,"prot_max":0.30,"fat_min":0.30,"fat_max":0.35,
                   "num_generations":120,"sol_per_pop":60,"mutation_percent_genes":8, "parent_selection_type":"random",
                   "keep_elitism":4,"w_prot":4.0,"w_fiber":3.5,"w_calcium":0.020,"w_iron":2.0,"w_vitc":0.6,"w_vita":0.005,
                   "w_cost100":7.0,"w_cost_porcion":4.0,"w_cost_lote":1.5},
    "75+ años":   {"carb_min":0.35,"carb_max":0.45,"prot_min":0.25,"prot_max":0.35,"fat_min":0.30,"fat_max":0.35,
                   "num_generations":130,"sol_per_pop":80,"mutation_percent_genes":6, "parent_selection_type":"rws",
                   "keep_elitism":5,"w_prot":5.0,"w_fiber":4.0,"w_calcium":0.025,"w_iron":2.5,"w_vitc":0.7,"w_vita":0.006,
                   "w_cost100":6.0,"w_cost_porcion":3.5,"w_cost_lote":1.0},
}

# ─────────────────────────────────────────────────────────────────────────────
# 3. FUNCIÓN DE FITNESS COMBINADA (3 costos)
# ─────────────────────────────────────────────────────────────────────────────
def compute_fitness(row, params):
    """Calcula el fitness de un plato dado los parámetros de su grupo de edad."""
    cals_p = row["Proteina_g"]    * 4
    cals_f = row["Lipidos_g"]     * 9
    cals_c = row["Carb_totales_g"]* 4
    total  = cals_p + cals_f + cals_c
    if total <= 0:
        return -9999.0, 0.0, 0.0, 0.0

    rp = cals_p / total
    rf = cals_f / total
    rc = cals_c / total

    def penalty(v, lo, hi):
        if v < lo: return (lo - v) ** 2
        if v > hi: return (v  - hi) ** 2
        return 0.0

    macro_pen = (penalty(rc, params["carb_min"], params["carb_max"]) +
                 penalty(rp, params["prot_min"], params["prot_max"]) +
                 penalty(rf, params["fat_min"],  params["fat_max"]))

    nutrition = (row["Proteina_g"]  * params["w_prot"]  +
                 row["Fibra_g"]     * params["w_fiber"]  +
                 row["Calcio_mg"]   * params["w_calcium"]+
                 row["Hierro_mg"]   * params["w_iron"]   +
                 row["VitC_mg"]     * params["w_vitc"]   +
                 row["VitA_ugER"]   * params["w_vita"])

    # Normalización de costos
    c100  = row["Costo_por_100g"]    / 25000.0
    cpor  = row["Precio_por_porcion"]/ 100000.0
    clote = row["Precio_total_lote"] / 1000000.0

    cost_pen = (params["w_cost100"]    * c100  +
                params["w_cost_porcion"]* cpor  +
                params["w_cost_lote"]  * clote)

    fitness = nutrition - 200.0 * macro_pen - cost_pen
    return fitness, rc * 100, rp * 100, rf * 100

# ─────────────────────────────────────────────────────────────────────────────
# 4. GENERAR RANKINGS
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 72)
print("  RANKING NUTRICIONAL COMPLETO POR RANGO DE EDAD")
print("=" * 72)

all_rankings = {}

for age_label, params in AGE_GROUPS.items():
    rows = []
    for _, row in df.iterrows():
        fitness, pct_c, pct_p, pct_f = compute_fitness(row, params)
        # Cumplimiento de rangos
        ok_c = params["carb_min"]*100 <= pct_c <= params["carb_max"]*100
        ok_p = params["prot_min"]*100 <= pct_p <= params["prot_max"]*100
        ok_f = params["fat_min"] *100 <= pct_f <= params["fat_max"] *100
        cumple = sum([ok_c, ok_p, ok_f])

        rows.append({
            "Platillo":         row["Platillo"],
            "Categoría":        row["Categoria"],
            "Fitness":          round(fitness, 3),
            "Energía (kcal)":   round(row["Energia_kcal"], 1),
            "Proteína (g)":     round(row["Proteina_g"], 2),
            "Carb. (g)":        round(row["Carb_totales_g"], 2),
            "Grasas (g)":       round(row["Lipidos_g"], 2),
            "% Carb":           round(pct_c, 1),
            "% Prot":           round(pct_p, 1),
            "% Grasa":          round(pct_f, 1),
            "Fibra (g)":        round(row["Fibra_g"], 2),
            "Calcio (mg)":      round(row["Calcio_mg"], 1),
            "Hierro (mg)":      round(row["Hierro_mg"], 2),
            "Vit.C (mg)":       round(row["VitC_mg"], 1),
            "Vit.A (μgER)":     round(row["VitA_ugER"], 1),
            "$/100g":           int(row["Costo_por_100g"]),
            "$/porción":        int(row["Precio_por_porcion"]),
            "$/lote":           int(row["Precio_total_lote"]),
            "Macros OK":        f"{cumple}/3",
            "✓Carb":            "✓" if ok_c else "✗",
            "✓Prot":            "✓" if ok_p else "✗",
            "✓Grasa":           "✓" if ok_f else "✗",
        })

    ranked = sorted(rows, key=lambda x: x["Fitness"], reverse=True)
    for i, r in enumerate(ranked):
        r["Posición"] = i + 1

    all_rankings[age_label] = ranked

    print(f"\n── {age_label}  (Sel:{params['parent_selection_type']} | Gen:{params['num_generations']} | Pop:{params['sol_per_pop']})")
    print(f"   Macros objetivo → Carb:{params['carb_min']*100:.0f}–{params['carb_max']*100:.0f}%  "
          f"Prot:{params['prot_min']*100:.0f}–{params['prot_max']*100:.0f}%  "
          f"Grasa:{params['fat_min']*100:.0f}–{params['fat_max']*100:.0f}%")
    for r in ranked[:5]:
        print(f"   {r['Posición']:2d}°  {r['Platillo']:<28s}  fit={r['Fitness']:7.2f}  "
              f"%C={r['% Carb']:4.1f}%  %P={r['% Prot']:4.1f}%  %G={r['% Grasa']:4.1f}%  "
              f"OK={r['Macros OK']}  $/p=${r['$/porción']:,}  $/100g=${r['$/100g']:,}")
    print(f"   {'─'*3}  ...{n_dishes - 5} platos más...")


# ─────────────────────────────────────────────────────────────────────────────
# 5. EXPORTAR A EXCEL — una hoja por grupo de edad + hoja resumen
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remover hoja vacía

# Colores por rango de edad
AGE_PALETTE = {
    "0–5 años":   ("1A5276", "D6EAF8"),
    "6–12 años":  ("1E8449", "D5F5E3"),
    "13–18 años": ("9A7D0A", "FCF3CF"),
    "19–35 años": ("884EA0", "F5EEF8"),
    "36–59 años": ("C0392B", "FDEDEC"),
    "60–75 años": ("1F618D", "EBF5FB"),
    "75+ años":   ("4A235A", "F4ECF7"),
}

MEDAL = {1: "🥇", 2: "🥈", 3: "🥉"}
thin = Side(style="thin", color="D0D0D0")
thick = Side(style="medium", color="888888")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
brd_top = Border(left=thin, right=thin, top=thick, bottom=thin)

def cs(cell, bold=False, bg=None, fc="000000", sz=9, ha="center", va="center", wrap=True):
    cell.font = Font(name="Calibri", bold=bold, color=fc, size=sz)
    if bg: cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    cell.border = brd

# ── Hoja resumen ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet("🏆 Resumen General")

RANK_COLS = 5  # top 5 platos por edad en resumen

title_val = "RANKING NUTRICIONAL — TOP 5 POR RANGO DE EDAD  |  Modelo: PyGAD con 3 métricas de costo"
ws_sum.merge_cells(f"A1:{get_column_letter(2 + RANK_COLS * 4)}1")
ws_sum["A1"] = title_val
cs(ws_sum["A1"], bold=True, bg="1C2833", fc="FFFFFF", sz=12)
ws_sum.row_dimensions[1].height = 24

sub = ("Fitness = Nutrición ponderada − Penalización macros − (w·$/100g + w·$/porción + w·$/lote).  "
       "Pesos y restricciones de macros varían por rango de edad.")
ws_sum.merge_cells(f"A2:{get_column_letter(2 + RANK_COLS * 4)}2")
ws_sum["A2"] = sub
cs(ws_sum["A2"], bg="2E4057", fc="BFC9CA", sz=8)
ws_sum.row_dimensions[2].height = 16

# Cabecera por grupo
headers_sum = ["Rango", "Objetivo macros"] + \
    [f"#{i}" for i in range(1, RANK_COLS + 1)] + \
    [f"Fitness #{i}" for i in range(1, RANK_COLS + 1)] + \
    [f"$/porción #{i}" for i in range(1, RANK_COLS + 1)] + \
    [f"$/100g #{i}" for i in range(1, RANK_COLS + 1)]

for ci, h in enumerate(headers_sum, 1):
    c = ws_sum.cell(row=3, column=ci, value=h)
    cs(c, bold=True, bg="2C3E50", fc="FFFFFF", sz=8)
ws_sum.row_dimensions[3].height = 22

for ri, (age_label, ranked) in enumerate(all_rankings.items(), 4):
    p = AGE_GROUPS[age_label]
    dark, light = AGE_PALETTE[age_label]

    ws_sum.cell(row=ri, column=1, value=age_label)
    cs(ws_sum.cell(row=ri, column=1), bold=True, bg=dark, fc="FFFFFF", sz=9, ha="left")

    macro_str = (f"C:{p['carb_min']*100:.0f}–{p['carb_max']*100:.0f}%  "
                 f"P:{p['prot_min']*100:.0f}–{p['prot_max']*100:.0f}%  "
                 f"G:{p['fat_min']*100:.0f}–{p['fat_max']*100:.0f}%")
    ws_sum.cell(row=ri, column=2, value=macro_str)
    cs(ws_sum.cell(row=ri, column=2), bg=light, sz=8, ha="left")

    for ki in range(RANK_COLS):
        r = ranked[ki]
        medal = MEDAL.get(ki + 1, "")
        # Nombre
        ws_sum.cell(row=ri, column=3 + ki, value=f"{medal} {r['Platillo']}")
        cs(ws_sum.cell(row=ri, column=3 + ki), bg=light, sz=8, ha="left", bold=(ki == 0))
        # Fitness
        ws_sum.cell(row=ri, column=3 + RANK_COLS + ki, value=round(r["Fitness"], 2))
        cs(ws_sum.cell(row=ri, column=3 + RANK_COLS + ki), bg=light, sz=8)
        # $/porción
        ws_sum.cell(row=ri, column=3 + 2*RANK_COLS + ki, value=r["$/porción"])
        cs(ws_sum.cell(row=ri, column=3 + 2*RANK_COLS + ki), bg=light, sz=8)
        ws_sum.cell(row=ri, column=3 + 2*RANK_COLS + ki).number_format = '$#,##0'
        # $/100g
        ws_sum.cell(row=ri, column=3 + 3*RANK_COLS + ki, value=r["$/100g"])
        cs(ws_sum.cell(row=ri, column=3 + 3*RANK_COLS + ki), bg=light, sz=8)
        ws_sum.cell(row=ri, column=3 + 3*RANK_COLS + ki).number_format = '$#,##0'

    ws_sum.row_dimensions[ri].height = 20

# Anchos resumen
ws_sum.column_dimensions["A"].width = 13
ws_sum.column_dimensions["B"].width = 28
for ci in range(3, 3 + RANK_COLS * 4):
    ws_sum.column_dimensions[get_column_letter(ci)].width = 24 if ci < 3 + RANK_COLS else 12
ws_sum.freeze_panes = "A4"


# ── Hoja detallada por grupo de edad ─────────────────────────────────────────
DET_COLS = [
    ("Posición",      "Pos.",        6),
    ("Platillo",      "Platillo",    26),
    ("Categoría",     "Categoría",   14),
    ("Fitness",       "Fitness",     9),
    ("Macros OK",     "Macros\nOK",  7),
    ("✓Carb",         "Carb\n✓✗",    6),
    ("✓Prot",         "Prot\n✓✗",    6),
    ("✓Grasa",        "Grasa\n✓✗",   6),
    ("% Carb",        "% Carb",      7),
    ("% Prot",        "% Prot",      7),
    ("% Grasa",       "% Grasa",     7),
    ("Energía (kcal)","Energía\n(kcal)",8),
    ("Proteína (g)",  "Proteína\n(g)",7),
    ("Carb. (g)",     "Carb.\n(g)",  7),
    ("Grasas (g)",    "Grasas\n(g)", 7),
    ("Fibra (g)",     "Fibra\n(g)",  7),
    ("Calcio (mg)",   "Calcio\n(mg)",8),
    ("Hierro (mg)",   "Hierro\n(mg)",8),
    ("Vit.C (mg)",    "Vit.C\n(mg)", 7),
    ("Vit.A (μgER)",  "Vit.A\n(μgER)",8),
    ("$/100g",        "$/100g",      10),
    ("$/porción",     "$/porción",   11),
    ("$/lote",        "$/lote",      12),
]

for age_label, ranked in all_rankings.items():
    p = AGE_GROUPS[age_label]
    dark, light = AGE_PALETTE[age_label]
    safe_name = age_label.replace("–", "-").replace("+", "+")
    ws = wb.create_sheet(safe_name)

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(DET_COLS))}1")
    ws["A1"] = f"RANKING — {age_label}  |  Carbohidratos: {p['carb_min']*100:.0f}–{p['carb_max']*100:.0f}%   Proteínas: {p['prot_min']*100:.0f}–{p['prot_max']*100:.0f}%   Grasas: {p['fat_min']*100:.0f}–{p['fat_max']*100:.0f}%"
    cs(ws["A1"], bold=True, bg=dark, fc="FFFFFF", sz=11)
    ws.row_dimensions[1].height = 22

    # Subtítulo modelo GA
    ws.merge_cells(f"A2:{get_column_letter(len(DET_COLS))}2")
    ws["A2"] = (f"Modelo GA: Selección={p['parent_selection_type'].upper()} | "
                f"Generaciones={p['num_generations']} | Población={p['sol_per_pop']} | "
                f"Mutación={p['mutation_percent_genes']}% | Élite={p['keep_elitism']}  ·  "
                f"Pesos fitness → Proteína×{p['w_prot']}  Fibra×{p['w_fiber']}  "
                f"Calcio×{p['w_calcium']}  Hierro×{p['w_iron']}  VitC×{p['w_vitc']}  VitA×{p['w_vita']}  "
                f"$/100g×{p['w_cost100']}  $/porción×{p['w_cost_porcion']}  $/lote×{p['w_cost_lote']}")
    cs(ws["A2"], bg="273746", fc="AEB6BF", sz=7, ha="left")
    ws.row_dimensions[2].height = 14

    # Cabeceras
    for ci, (key, label, _) in enumerate(DET_COLS, 1):
        c = ws.cell(row=3, column=ci, value=label)
        cs(c, bold=True, bg="2C3E50", fc="FFFFFF", sz=8)
    ws.row_dimensions[3].height = 30

    # Datos
    for r in ranked:
        row_num = r["Posición"] + 3
        pos = r["Posición"]

        # Color de fondo por posición
        if pos == 1:   bg = "FDFBCE"  # oro suave
        elif pos == 2: bg = "F2F3F4"  # plata suave
        elif pos == 3: bg = "FDEBD0"  # bronce suave
        elif r["Macros OK"] == "3/3": bg = "EBF5FB"  # cumple todo
        else: bg = "FFFFFF"

        for ci, (key, label, _) in enumerate(DET_COLS, 1):
            val = r[key]
            c = ws.cell(row=row_num, column=ci, value=val)
            ha = "left" if key in ("Platillo", "Categoría") else "center"
            bold = (pos <= 3 and key in ("Posición", "Platillo", "Fitness"))
            cs(c, bold=bold, bg=bg, sz=9, ha=ha)

            # Formato monetario
            if key == "$/100g":   c.number_format = '$#,##0'
            if key == "$/porción":c.number_format = '$#,##0'
            if key == "$/lote":   c.number_format = '$#,##0'
            if key == "Fitness":  c.number_format = '0.000'

            # Color de checks
            if key in ("✓Carb","✓Prot","✓Grasa"):
                c.font = Font(name="Calibri", bold=True, size=10,
                              color="1E8449" if val == "✓" else "C0392B")

        ws.row_dimensions[row_num].height = 16

    # Anchos
    for ci, (key, label, w) in enumerate(DET_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Formato condicional barras de datos para Fitness
    fit_range = f"D4:D{n_dishes + 3}"
    ws.conditional_formatting.add(fit_range,
        DataBarRule(start_type="min", end_type="max",
                    color=dark.replace("#","")))

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(DET_COLS))}3"


# ── Hoja comparativa de costos ────────────────────────────────────────────────
ws_cost = wb.create_sheet("💰 Análisis de Costos")

ws_cost.merge_cells("A1:H1")
ws_cost["A1"] = "ANÁLISIS COMPARATIVO DE COSTOS — Todos los platos"
cs(ws_cost["A1"], bold=True, bg="1C2833", fc="FFFFFF", sz=12)
ws_cost.row_dimensions[1].height = 22

cost_headers = ["Platillo","Categoría","$/100g (COP)","$/porción (COP)","$/lote (COP)",
                "Ranking $/100g","Ranking $/porción","Ranking $/lote"]
for ci, h in enumerate(cost_headers, 1):
    c = ws_cost.cell(row=2, column=ci, value=h)
    cs(c, bold=True, bg="2C3E50", fc="FFFFFF", sz=9)
ws_cost.row_dimensions[2].height = 20

df_cost = df[["Platillo","Categoria","Costo_por_100g","Precio_por_porcion","Precio_total_lote"]].copy()
df_cost["rank_100g"]   = df_cost["Costo_por_100g"].rank(ascending=True).astype(int)
df_cost["rank_porcion"]= df_cost["Precio_por_porcion"].rank(ascending=True).astype(int)
df_cost["rank_lote"]   = df_cost["Precio_total_lote"].rank(ascending=True).astype(int)
df_cost = df_cost.sort_values("Costo_por_100g")

for ri, (_, row) in enumerate(df_cost.iterrows(), 3):
    bg = "E8F8F5" if row["rank_100g"] <= 5 else ("FDFEFE" if ri % 2 == 0 else "FAFAFA")
    vals = [row["Platillo"], row["Categoria"],
            int(row["Costo_por_100g"]), int(row["Precio_por_porcion"]),
            int(row["Precio_total_lote"]),
            int(row["rank_100g"]), int(row["rank_porcion"]), int(row["rank_lote"])]
    for ci, v in enumerate(vals, 1):
        c = ws_cost.cell(row=ri, column=ci, value=v)
        ha = "left" if ci <= 2 else "center"
        cs(c, bg=bg, sz=9, ha=ha)
        if ci in [3, 4, 5]:
            c.number_format = '$#,##0'

    ws_cost.row_dimensions[ri].height = 16

ws_cost.column_dimensions["A"].width = 26
ws_cost.column_dimensions["B"].width = 16
for ci in range(3, 9):
    ws_cost.column_dimensions[get_column_letter(ci)].width = 16

ws_cost.freeze_panes = "A3"
ws_cost.auto_filter.ref = "A2:H2"

# Guardar
output_path = r"C:\Users\nikol\Documents\Ranking_Nutricional_Completo_PyGAD.xlsx"
wb.save(output_path)
print(f"\n✅ Ranking exportado: {output_path}")
print(f"   Hojas: Resumen General + {len(AGE_GROUPS)} hojas por edad + Análisis de Costos")